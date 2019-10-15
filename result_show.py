import matplotlib.pyplot as plt
from openpyxl import load_workbook

result_workbook = load_workbook('./paper/LJ/实验.xlsx')
result_sheetnames = result_workbook.get_sheet_names()
# wiki_sheet = result_workbook.get_sheet_by_name('Wiki_entity')
# wiki_sheet = result_workbook.get_sheet_by_name('Wiki_relation')

yago_sheet = result_workbook.get_sheet_by_name('Yago_entity')
yago_sheet = result_workbook.get_sheet_by_name('Yago_relation')


funcNames = ['TransE', 'TransH', 't-TransE', 'HyTE', 'RE-NET', 'T2LP']
# resultNames = ['Hits@1', 'Hits@3', 'Hits@5', 'Hits@7', 'Hits@10']
resultNames = ['Hits@1', 'Hits@3', 'Hits@5'] # just for yago relation

lineColor = ['b', 'g', 'r', 'c', 'm', 'k']
spotStyle = ['o-b', 'v-g', '^-r', 's-c', '*-m', 'D-k']

i = 0
# x = [1, 2, 3, 4, 5]
x = [1, 2, 3]

y = [0, 10, 20, 30, 40, 50, 60, 70, 80]
for result_row in range(2, 2+len(funcNames)):
    results = []
    for result_col in range(3, 3+len(resultNames)):
        # results.append(wiki_sheet.cell(row=result_row, column=result_col).value)
        results.append(yago_sheet.cell(row=result_row, column=result_col).value)
    print(results)
    plt.plot(x, results, lineColor[i], linewidth=1)
    plt.plot(x, results, spotStyle[i], linewidth=1, label=funcNames[i])
    i += 1

# x_labels = ['1', '3', '5', '7', '10']
x_labels = ['1', '3', '5']

y_labels = ['0', '10', '20', '30', '40', '50', '60', '70', '80']
# plt.title('Wiki Hits@K')
plt.xlabel('K')
plt.ylabel('Hits@K')

#
# fmt = '[marker][line][color]': https://matplotlib.org/3.1.1/api/_as_gen/matplotlib.pyplot.plot.html
# color:bgrcmykw
#
# plt.plot(x1, y1, 'r', linewidth=1, label='TransE')
# plt.plot(x1, y1, 'or', linewidth=5)
#
# plt.plot(x2, y2, 'g', linewidth=1, label='TransH')
# plt.plot(x2, y2, '^g', linewidth=5)
#
# plt.plot(x3, y3, 'r', linewidth=1, label='HyTE')
# plt.plot(x3, y3, '*r', linewidth=5)
#
# plt.plot(x4, y4, 'c', linewidth=1, label='T2LP')
# plt.plot(x4, y4, 'sc', linewidth=5)

plt.xticks(x, x_labels, rotation=0)
plt.yticks(y, y_labels, rotation=0)

plt.legend()
# plt.grid()
plt.show()